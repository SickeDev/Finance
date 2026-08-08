"""Importação em planilhas Excel (.xlsx).

Cada aba da planilha corresponde a uma coleção do sistema (accounts, cards,
card_purchases, boxes, investments, financings, recurring, transactions,
transfers, networth). A primeira linha é o cabeçalho com os nomes dos campos;
cada linha seguinte é um registro. Abas e campos aceitam português
(ex.: aba "Contas", coluna "Saldo").

O resultado retornado tem o mesmo formato de parse_import_xml (app/xml_utils.py),
então a importação reutiliza as mesmas regras de deduplicação e resolução de
referências (ex.: account_id pode ser o id ou o nome da conta).
"""

import datetime
import io
import re
import unicodedata

from openpyxl import load_workbook

from . import config

SHEET_ALIASES = {
    "accounts": "accounts",
    "contas": "accounts",
    "cards": "cards",
    "cartoes": "cards",
    "card_purchases": "card_purchases",
    "compras": "card_purchases",
    "boxes": "boxes",
    "caixinhas": "boxes",
    "investments": "investments",
    "investimentos": "investments",
    "financings": "financings",
    "financiamentos": "financings",
    "recurring": "recurring",
    "recorrentes": "recurring",
    "transactions": "transactions",
    "transacoes": "transactions",
    "transfers": "transfers",
    "transferencias": "transfers",
    "movimentacoes": "transfers",
    "payments": "payments",
    "pagamentos": "payments",
    "networth": "networth",
    "patrimonio": "networth",
    "settings": "settings",
    "configuracoes": "settings",
}

HEADER_ALIASES = {
    "id": "id",
    "nome": "name",
    "name": "name",
    "instituicao": "institution",
    "tipo": "type",
    "saldo": "balance",
    "descricao": "description",
    "categoria": "category",
    "valor": "amount",
    "data": "date",
    "limite": "limit",
    "dia_de_vencimento": "due_day",
    "dia_vencimento": "due_day",
    "bandeira": "brand",
    "parcelas": "installments",
    "parcelas_total": "installments_total",
    "total_de_parcelas": "installments_total",
    "valor_mensal": "monthly_value",
    "total": "total",
    "pago": "paid",
    "meta": "target",
    "ticker": "ticker",
    "quantidade": "quantity",
    "preco_medio": "avg_price",
    "preco_atual": "current_price",
    "frequencia": "frequency",
    "conta": "account_id",
    "conta_id": "account_id",
    "cartao": "card_id",
    "cartao_id": "card_id",
    "ativo": "active",
    "metodo": "method",
    "de_tipo": "from_type",
    "de": "from_id",
    "de_id": "from_id",
    "para_tipo": "to_type",
    "para": "to_id",
    "para_id": "to_id",
    "reserva_emergencia": "is_emergency",
}

THOUSANDS_RE = re.compile(r"^\d{1,3}(\.\d{3})+(,\d+)?$")
DATE_FORMATS = ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y")


def _norm(value):
    """Minúsculo, sem acentos e espaços viram underline (ex.: 'Dia do Vencimento')."""
    if not isinstance(value, str):
        value = str(value)
    value = value.strip().lower()
    value = "".join(c for c in unicodedata.normalize("NFD", value) if unicodedata.category(c) != "Mn")
    return value.replace(" ", "_").replace("-", "_")


def _field_name(header):
    return HEADER_ALIASES.get(_norm(header), _norm(header))


def _to_number(s):
    s = s.strip().replace("\xa0", "").replace(" ", "")
    if THOUSANDS_RE.match(s):
        if "," in s:
            return float(s.replace(".", "").replace(",", "."))
        return int(s.replace(".", ""))
    try:
        if "," in s:
            return float(s.replace(",", "."))
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return None


def _normalize_date(s):
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _cell_to_value(cell):
    if cell is None:
        return None
    if isinstance(cell, datetime.datetime):
        return cell.date().isoformat()
    if isinstance(cell, datetime.date):
        return cell.isoformat()
    if isinstance(cell, bool):
        return cell
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s:
        return None
    low = s.lower()
    if low in ("true", "false"):
        return low == "true"
    date_val = _normalize_date(s)
    if date_val is not None:
        return date_val
    num = _to_number(s)
    if num is not None:
        return num
    return s


def parse_import_xlsx(xlsx_bytes):
    """Lê um arquivo .xlsx e devolve {coleção: [docs]} no mesmo formato do XML."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        result = {}
        for sheet in wb.worksheets:
            col = SHEET_ALIASES.get(_norm(sheet.title))
            if col is None or col not in config.COLLECTIONS:
                continue
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            fields = []
            for i, h in enumerate(header):
                name = _field_name(h) if h is not None else ""
                if name:
                    fields.append((i, name))
            if not fields:
                continue
            items = []
            for row in rows:
                doc = {}
                for i, name in fields:
                    if i >= len(row):
                        continue
                    val = _cell_to_value(row[i])
                    if val is None:
                        continue
                    doc[name] = val
                if doc:
                    items.append(doc)
            if items:
                result[col] = items
        return result
    finally:
        wb.close()
